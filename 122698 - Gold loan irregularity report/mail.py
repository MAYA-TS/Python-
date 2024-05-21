import os
import pandas as pd
import cx_Oracle
import openpyxl
from email.message import EmailMessage
import smtplib
import pandas as pd
import cx_Oracle
import openpyxl
from email.message import EmailMessage
import pandas as pd
from openpyxl.styles import PatternFill,Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import smtplib
# from datetime import datetime, timedelta


# yesterday = datetime.now() - timedelta(days=1)
# yesterday_date = yesterday.strftime('%d-%m-%Y')

conn = cx_Oracle.connect("kpmg", "Asd$1234", "HISTDB")
print("Oracle database connected")


df1=pd.read_sql( """select m.*, c.emp_name irregularity_entered_empname, d.post_name
  from (select a.zonal_name zone,
               a.region_name region,
               a.inventory inventory_number,
               a.Irregularity,
               (count(verified_by) over(partition by inventory))Number_of_verifications_in_that_inventory,
               b.verified_by,
               b.verification_dt,
               a.loss loss_amount,
               a.tra_dt irregularity_entered_date,
               a.branch_name,
               a.cust_id customer_id,
               a.pledge_no,
               a.pledge_val pledge_amount,
               a.gross_weight weight,
               b.tra_dt inventory_creation_date,
               a.reportedby irregularity_entered_empcode,
               b.sticker_no from
        TABLEAU_AUDIT_IRR_RPT a
          left outer join (select a.pledge_no,
                                 a.sticker_no,
                                 trunc(a.tra_dt) verification_dt,
                                 a.usr verified_by,
                                 b.inv_id,
                                 b.tra_dt
                            from mana0809.goldloan_sticker_mst@uatr_backup2 a
                            left outer join (select inv_id, plgno, tra_dt
                                              from mana0809.tbl_gln_inventory_master@uatr_backup2) b
                              on a.pledge_no = b.plgno
                           where --post_id in (-252, 821, 309, 248)
                             --and
                              a.status = 2
                             and trunc(a.tra_dt) between
                                 trunc(add_months(trunc(sysdate), -12), 'mm') and
                                 trunc(sysdate - 1)) b
            on a.inventory = b.inv_id
         where trunc(a.tra_dt) between add_months(trunc(sysdate), -12) and
               trunc(sysdate - 1)
           and verification_dt < trunc(a.tra_dt)) m
  left outer join (select emp_code, emp_name, post_id
                     from mana0809.emp_master@uatr_backup2) c
    on m.irregularity_entered_empcode = c.emp_code
  left outer join (select a.emp_code, a.emp_name, a.post_id, b.post_name
                     from mana0809.emp_master@uatr_backup2 a,
                          mana0809.post_mst@uatr_backup2   b
                    where a.post_id = b.post_id) d
    on m.verified_by = d.emp_code""",con=conn)




writer=pd.ExcelWriter("Gold Loan Irregularity Report.xlsx",engine="openpyxl")
df1.to_excel(writer,sheet_name="Report",index=False)

print("saved as excel")

writer.save()

print("Excels Downloaded")


workbook = openpyxl.load_workbook(r"C:\Users\412919\Desktop\CRF\122698 - Gold loan irregularity report\Gold Loan Irregularity Report.xlsx")
sheet_names = workbook.sheetnames
heading_color =  '161a30'    #'73BCC5'#'8080ff'  # Red color
body_color = 'abd2e0'  # Green color
border_style = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin'))


for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color='161a30', end_color='161a30', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    body_fill = PatternFill(start_color=body_color, end_color=body_color, fill_type='solid')
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = body_fill
            
    for column in sheet.columns:
        non_empty_values = [cell.value for cell in column if cell.value]
        if non_empty_values:
            max_length = max(len(str(value)) for value in non_empty_values)
            column_letter = get_column_letter(column[0].column)
            adjusted_width = (max_length + 2) * 1.2  # Adjust the width as desired
            sheet.column_dimensions[column_letter].width = adjusted_width
    for row in sheet.rows:
        max_height = max(str(cell.value).count('\n') + 1 for cell in row if cell.value)
        row_number = row[0].row
        adjusted_height = max_height * 17 # Adjust the height as desired
        sheet.row_dimensions[row_number].height = adjusted_height
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border_style
workbook.save(r"C:\Users\412919\Desktop\CRF\122698 - Gold loan irregularity report\Gold Loan Irregularity Report.xlsx")


s = smtplib.SMTP(host='smtp.office365.com', port=587)
s.starttls()


s.login('internalaudit1@manappuram.com','AB@123ad')
#s.login('iotautomation@manappuram.com', 'ybjmxbfdyzkdnjtw')
msg = EmailMessage()


print("Ready for mailing")



subject = f'Gold Loan Irregularity Report for the previous 12 months'
msg['Subject'] = subject


#msg['From'] = 'IOT <iotautomation@manappuram.com>'
msg['From'] = 'INTERNAL AUDIT <internalaudit1@manappuram.com>'
msg['To'] = 'Branch Audit <branchaudit@manappuram.com>','Audit Research Wing<researchwing@manappuram.com>','MAFIL BRANCH AUDIT<internalaudit@manappuram.com>','MAFIL HO Audit<hoaudit@manappuram.com>'
#msg['To'] = 'SUGANYA DHANASEKARAN<ittesting16@manappuram.com>'
msg['Cc'] = 'RIJU P<gmaudit@manappuram.com>','sreelakshmi p satheesh <itprogramer20@manappuram.com>','MAYA T S<iotsupport7@manappuram.com>','ATHUL A<iotsupport10@manappuram.com>','DIBIN T R<iotsupport14@manappuram.com>','DINESH<itprogramer40@manappuram.com>'


# msg['To'] = 'Audit Research Wing<researchwing@manappuram.com>', 'RIJU P<gmaudit@manappuram.com>', 'Branch Audit <branchaudit@manappuram.com>', 'MAFIL BRANCH AUDIT<internalaudit@manappuram.com>', 'LAXMAN TAGGINAVAR <headresearchwing@manappuram.com>'
# msg['Cc'] = 'ASWIN DILEEP <iotsupport13@manappuram.com>','DEVIKA SUBASH<iotsupport16@manappuram.com>'

with open(r"C:\Users\412919\Desktop\CRF\122698 - Gold loan irregularity report\Gold Loan Irregularity Report.xlsx", 'rb') as ra:
    attachment = ra.read()
msg.add_related(attachment, maintype='application', subtype='xlsx', filename='Gold Loan Irregularity Report.xlsx')


msg.add_alternative("""\
    <html>
        <body>
            <p><i>Dear Sir/Madam,</i></p>
            <p> </p>
            <p><i> Kindly find the attachment for Gold loan irregularity report for the previous 12months and a daily report.</i></p>
            <p></p>
            <p> <i>Thanks & Regards,<br>
                ( This is an autogenerated mail )<br>
        R&D <br>
            </i></p>
        </body>
    </html>
    """ ,subtype='html')
s.send_message(msg)


print("Mail sent")

conn.close()