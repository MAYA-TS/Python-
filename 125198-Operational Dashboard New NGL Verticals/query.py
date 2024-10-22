from ctypes import alignment
import xlsxwriter
import os
import pandas as pd
import cx_Oracle
import openpyxl
from pandas import ExcelWriter
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from openpyxl.styles import Alignment

from email.message import EmailMessage
from email.utils import make_msgid
import mimetypes
import smtplib


conn = cx_Oracle.connect("kpmg", "Asd$1234", "HISTDB")
print("Oracle database connected")

#CKYC
df1=pd.read_sql("""select product_name,total_live_account, file_count, ckyc_id_created,
 upload_pending, total_ckyc_pending, within_tolerance, above_tolerance 
   from tableau_newngl_ckyc 
""",con=conn)

#INSURANCE
df2=pd.read_sql("""select PRODUCT_NAME,
       TOTAL_INSURANCE_PENDING,
       WITHIN_TOLERANCE,
       ABOVE_TOLERANCE
  from tableau_newngl_INSURANCE
""",con=conn)

# NACH
df3=pd.read_sql("""select PRODUCT_NAME,
       TOTAL_FILES_FOR_NACH,
       NACH_ACTIVATED,
       TOTAL_PENDING,
       WITHIN_TOLERANCE,
       ABOVE_TOLERANCE
  from tableau_newngl_nach
""",con=conn)


# QUERY
df4=pd.read_sql("""select PRODUCT_NAME,
       FILES_WITH_PENDING_QUERY,
       WITHIN_TOLERANCE,
       ABOVE_TOLERANCE
  from tableau_newngl_query
""",con=conn)

# FILE_MOVEMENT
df5=pd.read_sql("""select DEP_NAME,
       TOTAL_FILES,
       FILES_TO_BE_TRANSFERRED_TO_HO,
       FILES_RECEIVED_HO,
       COMPLETION_PER,
       PENDING,
       PENDING_WITHIN_TOLERANCE,
       PENDING_ABOVE_TOLERANCE,
       FILE_UPLOAD_IN_SYSTEM_PENDING,
       WITHIN_TOLERANCE,
       ABOVE_TOLERANCE,
       VERIFICATIONCOMPLETION_PER
  from tableau_newngl_FILE_MOVEMENT
""",con=conn)

print("Excels downloaded")

writer = pd.ExcelWriter("Operational_Dashboard.xlsx", engine="openpyxl")

df1.to_excel(writer, sheet_name="CKYC", index=False)
df2.to_excel(writer, sheet_name="INSURANCE", index=False)
df3.to_excel(writer, sheet_name="NACH", index=False)
df4.to_excel(writer, sheet_name="QUERY", index=False)
df5.to_excel(writer, sheet_name="FILE_MOVEMENT", index=False)

writer.close()

# loading the excel for editing
wb = load_workbook("Operational_Dashboard.xlsx")
# eduting

# CKYC
def CKYC():
  ckyc_sheet = wb["CKYC"]
  ckyc_sheet.insert_rows(1)
  ckyc_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
  ckyc_sheet["A1"] = "CKYC (TAT: 15)"
  ckyc_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
CKYC()

# Insurance
def insurance():
  ckyc_sheet = wb["INSURANCE"]
  ckyc_sheet.insert_rows(1)
  ckyc_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
  ckyc_sheet["A1"] = "INSURANCE"
  ckyc_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
insurance()

# NACH
def NACH():
  ckyc_sheet = wb["NACH"]
  ckyc_sheet.insert_rows(1)
  ckyc_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
  ckyc_sheet["A1"] = "NACH"
  ckyc_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
NACH()

# query
def QUERY():
  ckyc_sheet = wb["QUERY"]
  ckyc_sheet.insert_rows(1)
  ckyc_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
  ckyc_sheet["A1"] = "QUERY (TAT  30)"
  ckyc_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
QUERY()

# FILE_MOVEMENT
def FILE():
  ckyc_sheet = wb["FILE_MOVEMENT"]
  ckyc_sheet.insert_rows(1)
  ckyc_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
  ckyc_sheet["A1"] = "FILE MOVEMENT TO HO (TAT : 10 )"
  ckyc_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
FILE()

wb.save("Operational_Dashboard.xlsx")
print("Text added, merged, and Excel saved.")




# #mail

# s = smtplib.SMTP(host='smtp.office365.com', port=587)
# s.starttls()

# s.login('iotautomation@manappuram.com', 'ybjmxbfdyzkdnjtw')
# # s.login('internalaudit1@manappuram.com','AB@123ad')
# msg = EmailMessage()

# print("Ready for mailing")
# subject = f'Operation Dashboard'
# msg['Subject'] = subject

        
# msg['From'] = 'IOT <iotautomation@manappuram.com>'

# # msg['To']='Leya S Jaya<ittesting25@manappuram.com>'
# msg['Cc'] ='MAYA T S<iotsupport7@manappuram.com>','Prajith V N<am7dataservice@manappuram.com>','Saneesh P B<dataservice32@manappuram.com>','DIBIN T R<iotsupport14@manappuram.com>'


# #C:\Users\412919\Desktop\CRF\123742-PAN VERIFICATION NEW DASHBOARD
# # F:\\CRF\\125198-Operational Dashboard New NGL Verticals

# with open(r"F:\\CRF\\125198-Operational Dashboard New NGL Verticals\\Operational_Dashboard.xlsx", 'rb') as ra:
#         attachment = ra.read()
             
# msg.add_related(attachment, maintype='application', subtype='xlsx', filename='Operational_Dashboard.xlsx')  

# msg.add_alternative("""   
#                             <html>
#         <body>
#             <p><i>Dear Sir/Madam,</i></p>
#             <p> </p>
#             <p><i> Kindly find the attachment of CRF for Operation Dashboard New NGL Verticals(CRF Id:125198).Please check and verify.</i></p>
#             <p></p>
#             <p> <i>Thanks & Regards,<br>
#                 ( This is an autogenerated mail )<br>
#         R&D <br>
#             </i></p>
#         </body>
#     </html>
#     """ ,subtype='html')
        
# s.send_message(msg)

# print("Mail sent")
